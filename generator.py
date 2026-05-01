"""
Генерация книги посещаемости (Docházka) из таблицы часов (Mzdy).
"""

import re
import calendar
import unicodedata
import openpyxl
from dataclasses import dataclass, field
from datetime import time, date, datetime
from openpyxl.styles import Font

CZECH_MONTHS = {
    1: "Leden", 2: "Únor", 3: "Březen", 4: "Duben",
    5: "Květen", 6: "Červen", 7: "Červenec", 8: "Srpen",
    9: "Září", 10: "Říjen", 11: "Listopad", 12: "Prosinec",
}


class DochazkaError(Exception):
    """Ошибка с понятным сообщением для пользователя."""
    pass


@dataclass
class Summary:
    month: int
    year: int
    employees: dict[str, float] = field(default_factory=dict)

    def format_text(self) -> str:
        month_name = CZECH_MONTHS.get(self.month, str(self.month))
        total = sum(self.employees.values())
        lines = [f"{month_name} {self.year}", ""]
        for name in sorted(self.employees):
            lines.append(f"  {name}: {self.employees[name]:g} ч.")
        lines.append(f"\nВсего: {total:g} ч.")
        return "\n".join(lines)


def _remove_diacritics(s: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )


def _normalize(s: str) -> str:
    return _remove_diacritics(s.strip()).lower()


WEEKDAY_OPEN = (6, 30)
WEEKEND_OPEN = (8, 0)


def _minutes_to_time(m: int) -> time:
    m = max(0, min(m, 23 * 60 + 59))
    return time(m // 60, m % 60)


def _get_opening_minutes(d: date) -> int:
    if d.weekday() >= 5:
        return WEEKEND_OPEN[0] * 60 + WEEKEND_OPEN[1]
    return WEEKDAY_OPEN[0] * 60 + WEEKDAY_OPEN[1]


def assign_shifts(emp_hours: dict[str, float], operating_hours: float,
                  day_date: date) -> dict[str, tuple[time, time]]:
    """
    Assign shift times for employees within cafe operating hours.

    Returns: {name: (arrival_time, departure_time)}
    """
    if not emp_hours:
        return {}

    open_min = _get_opening_minutes(day_date)
    close_min = open_min + int(operating_hours * 60)

    employees = sorted(emp_hours.keys())
    n = len(employees)
    shifts = {}

    if n == 1:
        emp = employees[0]
        start = open_min
        end = start + int(emp_hours[emp] * 60)
        shifts[emp] = (start, end)

    elif n == 2:
        emp_a, emp_b = employees[0], employees[1]
        a_start = open_min
        a_end = a_start + int(emp_hours[emp_a] * 60)
        b_end = close_min
        b_start = b_end - int(emp_hours[emp_b] * 60)
        shifts[emp_a] = (a_start, a_end)
        shifts[emp_b] = (b_start, b_end)

    else:
        current = open_min
        for emp in employees:
            start = current
            end = start + int(emp_hours[emp] * 60)
            shifts[emp] = (start, end)
            current = end
        last_emp = employees[-1]
        _, last_end = shifts[last_emp]
        if last_end > close_min:
            diff = last_end - close_min
            old_start = shifts[last_emp][0]
            shifts[last_emp] = (old_start - diff, close_min)

    return {
        emp: (_minutes_to_time(s), _minutes_to_time(e))
        for emp, (s, e) in shifts.items()
    }


def parse_mzdy(filepath: str):
    """
    Parse a Mzdy Excel file.

    Returns:
        month, year, employee_names, full_names, days

    Where:
        employee_names: list[str] — last names from headers
        full_names: dict[str, str] — {normalized_last_name: "First Last"}
        days: list[dict] — [{day, operating_hours, employee_hours: {name: hours}}]

    Raises DochazkaError with a user-friendly message on any problem.
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception:
        raise DochazkaError(
            "Не удалось прочитать файл. Убедись, что это файл Excel (.xlsx)."
        )

    # Month/year from filename
    match = re.search(r"(\d{1,2})[._](\d{4})", filepath)
    if match:
        month = int(match.group(1))
        year = int(match.group(2))
    else:
        raise DochazkaError(
            "Не удалось определить месяц и год из имени файла.\n"
            "Файл должен называться Mzdy_MM.YYYY.xlsx\n"
            "Например: Mzdy_03.2026.xlsx"
        )

    # Find the sheet with daily data (B1 contains "час")
    ws = None
    ws_names = None
    for sheet_name in wb.sheetnames:
        candidate = wb[sheet_name]
        b1 = candidate.cell(row=1, column=2).value
        if b1 and isinstance(b1, str) and "час" in b1.lower():
            ws = candidate
        else:
            if ws_names is None:
                ws_names = candidate

    if ws is None:
        raise DochazkaError(
            "Не нашёл лист с данными.\n"
            "В ячейке B1 первого листа должно быть слово «часов».\n"
            "Проверь, что файл заполнен по шаблону."
        )

    # Employee names from row 1, columns D+
    employee_names = []
    for col in range(4, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val and isinstance(val, str) and val.strip():
            employee_names.append(val.strip())

    if not employee_names:
        raise DochazkaError(
            "Не нашёл имена сотрудников в строке 1 (столбцы D и далее).\n"
            "Впиши фамилии сотрудников в первую строку, начиная со столбца D."
        )

    # Full names from second sheet (if present)
    full_names = {}
    if ws_names is not None:
        for row in range(1, ws_names.max_row + 1):
            for fc, lc in [(1, 2), (3, 4)]:
                first = ws_names.cell(row=row, column=fc).value
                last = ws_names.cell(row=row, column=lc).value
                if (first and last
                        and isinstance(first, str) and isinstance(last, str)
                        and first.lower() not in ("jmeno", "jméno", "celkem")
                        and last.lower() not in ("príjmení", "příjmení")):
                    key = _normalize(last)
                    if key not in full_names:
                        full_names[key] = f"{first.strip()} {last.strip()}"

    # Daily data (rows 5-35)
    days = []
    for row in range(5, 36):
        a_val = ws.cell(row=row, column=1).value
        if not a_val:
            continue

        m = re.match(r"(\d+)", str(a_val))
        if not m:
            continue
        day_num = int(m.group(1))

        b_val = ws.cell(row=row, column=2).value
        if not b_val:
            continue
        try:
            operating_hours = float(b_val)
        except (ValueError, TypeError):
            continue

        emp_hours = {}
        for i, name in enumerate(employee_names):
            val = ws.cell(row=row, column=4 + i).value
            if val is not None:
                try:
                    h = float(val)
                    if h > 0:
                        emp_hours[name] = h
                except (ValueError, TypeError):
                    pass

        if emp_hours:
            days.append({
                "day": day_num,
                "operating_hours": operating_hours,
                "employee_hours": emp_hours,
            })

    wb.close()

    if not days:
        raise DochazkaError(
            "Нет данных по дням (строки 5–35 пустые).\n"
            "Заполни часы работы кофейни (столбец B) и часы сотрудников (столбцы D+)."
        )

    return month, year, employee_names, full_names, days


def _time_to_excel(t: time) -> float:
    return (t.hour * 60 + t.minute) / 1440


def generate(input_path: str, output_path: str) -> Summary:
    """
    Parse Mzdy file, generate Docházka file.

    Returns Summary with stats.
    Raises DochazkaError on problems.
    """
    month, year, employee_names, full_names, days = parse_mzdy(input_path)

    num_days = calendar.monthrange(year, month)[1]
    month_name = CZECH_MONTHS.get(month, str(month))

    # Compute all shifts
    all_shifts = {}
    for day_data in days:
        d = day_data["day"]
        day_date = date(year, month, d)
        shifts = assign_shifts(
            day_data["employee_hours"],
            day_data["operating_hours"],
            day_date,
        )
        all_shifts[d] = shifts

    # Find active employees (with >0 hours in the month)
    active = {}
    for day_data in days:
        for name, hours in day_data["employee_hours"].items():
            if hours > 0:
                active.setdefault(name, 0)
                active[name] += hours

    # Generate workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for emp_name in sorted(active.keys()):
        norm = _normalize(emp_name)
        display_name = full_names.get(norm, emp_name)

        parts = display_name.split()
        if len(parts) >= 2:
            tab_name = f"{parts[0][0]}. {parts[1]}"
        else:
            tab_name = emp_name
        tab_name = tab_name[:31]

        ws = wb.create_sheet(title=tab_name)

        ws["A1"] = "Kniha docházky"
        ws["A1"].font = Font(bold=True, size=12)
        ws["B1"] = f"{month_name} {year}"
        ws["B1"].font = Font(size=11)
        ws["A3"] = display_name
        ws["A3"].font = Font(bold=True, size=11)

        headers = [
            "Datum", "Příchod", "Odchod", "Přestávka (min)",
            "Odpracováno (hod)", "Podpis zaměstnance",
            "Podpis zaměstnavatele", "Poznámka",
        ]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=i, value=h)
            cell.font = Font(bold=True)

        total_hours = 0
        row = 5
        for day_num in range(1, num_days + 1):
            day_dt = datetime(year, month, day_num)
            ws.cell(row=row, column=1, value=day_dt)
            ws.cell(row=row, column=1).number_format = "dd.mm.yyyy"

            if day_num in all_shifts and emp_name in all_shifts[day_num]:
                arrival, departure = all_shifts[day_num][emp_name]

                ws.cell(row=row, column=2, value=_time_to_excel(arrival))
                ws.cell(row=row, column=2).number_format = "h:mm"
                ws.cell(row=row, column=3, value=_time_to_excel(departure))
                ws.cell(row=row, column=3).number_format = "h:mm"

                for dd in days:
                    if dd["day"] == day_num and emp_name in dd["employee_hours"]:
                        hours = dd["employee_hours"][emp_name]
                        ws.cell(row=row, column=5, value=hours)
                        total_hours += hours
                        break

            row += 1

        ws.cell(row=row, column=4, value="Součet hodin:")
        ws.cell(row=row, column=4).font = Font(bold=True)
        ws.cell(row=row, column=5, value=total_hours)
        ws.cell(row=row, column=5).font = Font(bold=True)

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 22
        ws.column_dimensions["H"].width = 14

    wb.save(output_path)

    return Summary(month=month, year=year, employees=active)

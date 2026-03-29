import os
import tempfile
import openpyxl
from datetime import date, time
from generator import DochazkaError, Summary, parse_mzdy, assign_shifts


def test_dochazka_error_has_message():
    err = DochazkaError("Не нашёл лист с данными.")
    assert str(err) == "Не нашёл лист с данными."


def test_summary_fields():
    s = Summary(month=3, year=2026, employees={"Novak": 120.0, "Svoboda": 80.0})
    assert s.month == 3
    assert s.year == 2026
    assert s.employees == {"Novak": 120.0, "Svoboda": 80.0}


def test_summary_format_text():
    s = Summary(month=3, year=2026, employees={"Novak": 120.0, "Svoboda": 80.0})
    text = s.format_text()
    assert "Březen 2026" in text
    assert "Novak" in text
    assert "120" in text
    assert "Svoboda" in text
    assert "80" in text
    assert "200" in text  # total


def create_test_mzdy(filename="Mzdy_03.2026.xlsx", days=None, employees=None,
                      full_names=None, b1_value="кол-во часов"):
    """Create a test Mzdy Excel file. Returns path to the file."""
    if employees is None:
        employees = ["Novak", "Svoboda"]
    if days is None:
        days = {
            1: {"cafe": 12, "hours": {"Novak": 8, "Svoboda": 6}},
            2: {"cafe": 10, "hours": {"Novak": 5}},
        }

    tmp_dir = tempfile.mkdtemp()
    path = os.path.join(tmp_dir, filename)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hodiny"

    # Row 1: headers
    ws.cell(row=1, column=2, value=b1_value)
    for i, name in enumerate(employees):
        ws.cell(row=1, column=4 + i, value=name)

    # Rows 5-35: daily data
    for day_num, data in days.items():
        row = 4 + day_num
        ws.cell(row=row, column=1, value=str(day_num))
        ws.cell(row=row, column=2, value=data["cafe"])
        for i, emp in enumerate(employees):
            if emp in data.get("hours", {}):
                ws.cell(row=row, column=4 + i, value=data["hours"][emp])

    # Optional: second sheet with full names
    if full_names:
        ws2 = wb.create_sheet("Zaměstnanci")
        ws2.cell(row=1, column=1, value="Jméno")
        ws2.cell(row=1, column=2, value="Příjmení")
        for i, (first, last) in enumerate(full_names, 2):
            ws2.cell(row=i, column=1, value=first)
            ws2.cell(row=i, column=2, value=last)

    wb.save(path)
    wb.close()
    return path


def test_parse_mzdy_basic():
    path = create_test_mzdy()
    month, year, emp_names, full_names, days = parse_mzdy(path)
    assert month == 3
    assert year == 2026
    assert "Novak" in emp_names
    assert "Svoboda" in emp_names
    assert len(days) == 2
    assert days[0]["day"] == 1
    assert days[0]["operating_hours"] == 12
    assert days[0]["employee_hours"]["Novak"] == 8
    os.unlink(path)


def test_parse_mzdy_with_full_names():
    path = create_test_mzdy(full_names=[("Jan", "Novák"), ("Petra", "Svobodová")])
    month, year, emp_names, full_names, days = parse_mzdy(path)
    assert full_names["novak"] == "Jan Novák"
    assert full_names["svobodova"] == "Petra Svobodová"
    os.unlink(path)


def test_parse_mzdy_no_hours_sheet():
    path = create_test_mzdy(b1_value="something else")
    try:
        parse_mzdy(path)
        assert False, "Should have raised DochazkaError"
    except DochazkaError as e:
        assert "часов" in str(e)
    finally:
        os.unlink(path)


def test_parse_mzdy_no_employees():
    path = create_test_mzdy(employees=[])
    try:
        parse_mzdy(path)
        assert False, "Should have raised DochazkaError"
    except DochazkaError as e:
        assert "сотрудник" in str(e).lower()
    finally:
        os.unlink(path)


def test_parse_mzdy_no_data():
    path = create_test_mzdy(days={})
    try:
        parse_mzdy(path)
        assert False, "Should have raised DochazkaError"
    except DochazkaError as e:
        assert "данных" in str(e).lower()
    finally:
        os.unlink(path)


def test_assign_shifts_one_employee():
    # Monday, opens at 6:30
    shifts = assign_shifts({"Novak": 8}, 12, date(2026, 3, 2))
    assert shifts["Novak"] == (time(6, 30), time(14, 30))


def test_assign_shifts_two_employees():
    # Monday, opens 6:30, closes 18:30 (12h)
    shifts = assign_shifts({"Novak": 8, "Svoboda": 6}, 12, date(2026, 3, 2))
    # Novak opens: 6:30 - 14:30
    assert shifts["Novak"] == (time(6, 30), time(14, 30))
    # Svoboda closes: 12:30 - 18:30
    assert shifts["Svoboda"] == (time(12, 30), time(18, 30))


def test_assign_shifts_three_employees():
    # Monday, opens 6:30, closes 18:30 (12h)
    shifts = assign_shifts(
        {"Novak": 4, "Kral": 4, "Svoboda": 4}, 12, date(2026, 3, 2)
    )
    # Sequential from opening (sorted alphabetically: Kral, Novak, Svoboda)
    assert shifts["Kral"] == (time(6, 30), time(10, 30))
    assert shifts["Novak"] == (time(10, 30), time(14, 30))
    assert shifts["Svoboda"] == (time(14, 30), time(18, 30))


def test_assign_shifts_weekend():
    # Saturday, opens at 8:00
    shifts = assign_shifts({"Novak": 6}, 10, date(2026, 3, 7))
    assert shifts["Novak"] == (time(8, 0), time(14, 0))


def test_assign_shifts_empty():
    shifts = assign_shifts({}, 12, date(2026, 3, 2))
    assert shifts == {}
